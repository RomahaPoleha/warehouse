from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import EquipmentType, Equipment, RepairRequest
from django.utils import timezone
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth import login, logout, authenticate
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import datetime
from django.http import HttpResponse
from .forms import SubtractQuantityForm
from django.db import transaction

# @csrf_exempt
# def create_superuser(request):
#     # Проверяем, есть ли уже admin
#     if User.objects.filter(username='admin').exists():
#         return HttpResponse('''
#             <h1>✅ Admin уже существует!</h1>
#             <p>Теперь удали этот код из views.py</p>
#             <p><a href="/admin/">Перейти в админку</a></p>
#         ''')
#
#     # Создаём суперюзера
#     User.objects.create_superuser('admin', 'admin@example.com', 'NewPass2024!')
#     return HttpResponse('''
#         <h1>✅ Admin создан!</h1>
#         <p><b>Логин:</b> admin</p>
#         <p><b>Пароль:</b> NewPass2024!</p>
#         <p><a href="/admin/">Перейти в админку</a></p>
#         <p style="color: red;"><b>Теперь удали этот код из views.py!</b></p>
#     ''')

# Домашняя страница
@login_required
def home(request):
    return render(request, 'accounting/home.html')

@login_required
# Список ИК-Рамок
def ir_frames(request):
    query = request.GET.get('q','').strip()
    ir_frames = Equipment.objects.filter(equip_type=EquipmentType.IR_FRAMES)
    if query:
        ir_frames = ir_frames.filter(Q(name__icontains=query)| Q(model__icontains=query))
    return render(request, 'accounting/ir_frames_list.html', {'ir_frames': ir_frames, 'query':query,'has_query': bool(query)})

@login_required
# Список Фронтальных плат и плат подключения OPS
def front_and_connector(request):
    query = request.GET.get('q', '').strip()
    front_and_connector = Equipment.objects.filter(equip_type=EquipmentType.FRONT_CONNECTOR)
    if query:
        front_and_connector =front_and_connector.filter(Q(name__icontains=query)| Q(model__icontains=query)|Q(modification__icontains=query))
    return render(request, 'accounting/front_and_connector_list.html', {'front_and_connector': front_and_connector,'query':query, 'has_query': bool(query)})

@login_required
# Список устройст на базе андроид
def board_android(request):
    query = request.GET.get('q','').strip()
    board_androids = Equipment.objects.filter(equip_type=EquipmentType.ANDROID_BOARD)
    if query:
        board_androids = board_androids.filter(Q(name__icontains=query)| Q(model__icontains=query)|Q(modification__icontains=query))

    return render(request, 'accounting/board_android_list.html', {'board_androids': board_androids,'query':query, 'has_query': bool(query)})

@login_required
# Список устройст на базе Windows
def board_windows(request):
    query = request.GET.get("q","").strip()
    board_windows = Equipment.objects.filter(equip_type=EquipmentType.WINDOWS_BOARD)
    if query:
        board_windows = board_windows.filter(Q(name__icontains=query)| Q(model__icontains=query))
    return render(request, 'accounting/board_windows_list.html', {'board_windows': board_windows, 'query':query, 'has_query': bool(query)})

@login_required
# Список плат T-con
def tcon(request):
    query = request.GET.get('q','').strip()
    tcons = Equipment.objects.filter(equip_type=EquipmentType.TCON)
    if query:
        tcons = tcons.filter(Q(name__icontains=query)| Q(model__icontains=query))
    return render(request, 'accounting/tcon_list.html', {'tcons': tcons, 'query':query, 'has_query': bool(query) })

@login_required
# Список кабелей
def cable(request):
    query = request.GET.get('q', '').strip()
    cables= Equipment.objects.filter(equip_type=EquipmentType.CABLE)
    if query:
        cables = cables.filter(Q(name__icontains=query)| Q(model__icontains=query))
    return render(request, 'accounting/cable_list.html', {'cables': cables,'query':query,'has_query': bool(query)})

@login_required
# Список устройст на базе андроид
def videowall(request):
    query = request.GET.get('q','').strip()
    videowall_list = Equipment.objects.filter(equip_type=EquipmentType.VIDEOWALL)
    if query:
        videowall_list = videowall_list.filter(Q(name__icontains=query)| Q(model__icontains=query))

    return render(request, 'accounting/videowall_list.html', {'videowall_list': videowall_list,'query':query, 'has_query': bool(query)})


@staff_member_required
# Вывод всех запросов
def request_list(request):
    req_pending =  RepairRequest.objects.filter(status='pending')
    return render(request, 'accounting/request_list.html', {'req_pending': req_pending})

@staff_member_required
# Вывод истории выдачи
def history_list(request):
    req_history = RepairRequest.objects.filter(status__in=['confirmed', 'rejected'])  # Фильтр для 2х условий
    return render(request, 'accounting/history_list.html', {'req_history': req_history})

@login_required
# Логика оформления заявки для запроса оборудования
def request_item(request, pk):
    # 1. Находим оборудование
    req_item = get_object_or_404(Equipment, pk=pk)

    # Получение страницы возврата
    next_page = request.GET.get('next', "home")
    # 2. Если нажата кнопка запросить
    if request.method == "POST":
        try:
            quantity = int(request.POST.get("quantity", 0))
        except ValueError:
            messages.error(request, 'Некорректное количество')
            return redirect('board_android_list')

        # ✅ Проверка: количество больше 0
        if quantity <= 0:
            messages.error(request, 'Количество должно быть больше 0')
            return redirect('request_item', pk=pk)

        # ✅ Проверка: хватает ли на складе
        if quantity > req_item.quantity:
            messages.error(request, f"На складе только {req_item.quantity} шт.")
            return redirect('request_item', pk=pk)

        # 3. Создание заявки
        RepairRequest.objects.create(
            equipment=req_item,
            requesting_user=request.user,
            quantity_requested=quantity,
            status="pending"
        )

        messages.success(request, 'Заявка создана!')
        return redirect(next_page)

    return render(request, 'accounting/request_item.html', {
        'req_item': req_item,
        'next_page': next_page
    })

@staff_member_required
# Логика выдачи оборудования
def confirm_request(request, request_id):
    # Находим заявку
    req_item = get_object_or_404(RepairRequest, id=request_id)

    # Проверяем статус
    if req_item.status != "pending":
        messages.error(request,'Заявку уже обработали')
        return redirect('request_list')

    # Находим оборудование
    equipment = req_item.equipment

    # Проверяем остаток
    if equipment.quantity < req_item.quantity_requested:
        messages.error(request, f"Недостаточное количество. Осталось: {equipment.quantity}")
        return redirect ("request_list")

    # Списываем со склада
    equipment.quantity -= req_item.quantity_requested
    equipment.save()

    # Присваиваем статус
    req_item.status = 'confirmed'

    # Фиксируем время и дату выдачи
    req_item.confirmed_at = timezone.now()
    req_item.approved_user = request.user
    req_item.save()

    messages.success(request, f"Количество {req_item.quantity_requested} успешно выдано")
    return redirect ("request_list")

@staff_member_required
# Логика отказа о выдаче
def reject_application(request, req_id):
    reject = get_object_or_404(RepairRequest, id=req_id)

    # Проверка: заявку можно отклонить, только если она в статусе "Ожидает"
    if reject.status != "pending":
        messages.error(request, 'Заявку уже обработали')
        return redirect("request_list")

    # Заполняем данные об отклонении
    reject.status = "rejected"
    reject.approved_user = request.user      # ← Кто отклонил
    reject.confirmed_at = timezone.now()     # ← Когда отклонил
    reject.save()

    messages.success(request, f"Заявка #{reject.id} успешно отклонена")  # ← Добавлен request
    return redirect("request_list")


# Страница входа
def login_view(request):
    if request.user.is_authenticated:
        return redirect('home')

    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            messages.success(request, f'Добро пожаловать, {user.username}!')
            return redirect('home')
    else:
        form = AuthenticationForm()

    return render(request, 'accounting/login.html', {'form': form})

# Страница выхода
def logout_view(request):
    logout(request)
    messages.info(request,'Вы успешно вышли из системы')
    return redirect('login')

# Логика вывода заявок для пользователя
@login_required  # ← Только для авторизованных
def history_user(request):
    # Фильтруем заявки: только для текущего пользователя
    user_requests = RepairRequest.objects.filter(
        requesting_user=request.user
    ).select_related('equipment')  # ← Оптимизация: сразу подгружаем оборудование

    return render(request, "accounting/my_request.html", {"user_requests": user_requests})


@login_required
def export_equipment_to_excel(request):
    """Выгружает всё оборудование в Excel"""
    # Создаём книгу и лист
    wb = Workbook()
    ws = wb.active
    ws.title = "Оборудование"

    # Заголовки с форматированием
    headers = [
        'Название', 'Модель', 'Модификация',
        'Длина (см)', 'Количество', 'Ячейки',
        'Описание', 'Процессор'
    ]

    # Стиль заголовка
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Данные
    for row, equip in enumerate(Equipment.objects.all().order_by('id'), start=2):
            ws.cell(row=row, column=1, value=equip.name)
            ws.cell(row=row, column=2, value=equip.model)
            if not equip.modification:
                equip.modification = 'x'
                ws.cell(row=row, column=3, value=equip.modification)
            else:
                ws.cell(row=row, column=3, value=equip.modification)
            if not equip.length_cm:
                equip.length_cm = 'x'
                ws.cell(row=row, column=4, value=equip.length_cm)
            else:
                ws.cell(row=row, column=4, value=equip.length_cm)
            ws.cell(row=row, column=5, value=equip.quantity)
            ws.cell(row=row, column=6, value=equip.cells)
            if not equip.description:
                equip.description = 'x'
                ws.cell(row=row, column=7, value=equip.description)
            else:
                ws.cell(row=row, column=7, value=equip.description)
            if not equip.cpu:
                equip.cpu = 'x'
                ws.cell(row=row, column=8, value=equip.cpu)
            else:
                ws.cell(row=row, column=8, value=equip.cpu)


    # Автоширина колонок
    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

    # Ответ с файлом
    filename = f'equipment_{datetime.date.today()}.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response

# Уменьшение количества
@transaction.atomic
def subtract_quantity_view(request, product_id):
    if request.method == 'POST':
        form = SubtractQuantityForm(request.POST)
        if form.is_valid():
            amount = form.cleaned_data['amount']

            # Блокируем строку на время транзакции
            product = Equipment.objects.select_for_update().get(pk=product_id)

            if product.quantity >= amount:
                product.quantity -= amount
                product.save()

                messages.success(request, f"Успешно списано {amount} шт.")

                # Возврат на страницу, откуда пришли
                next_url = request.GET.get('next', 'home')
                return redirect(next_url)
            else:
                messages.error(request, f"Недостаточно товара. Доступно: {product.quantity} шт.")
    else:
        # Для GET-запроса — обычный get_object_or_404 (без транзакции)
        product = get_object_or_404(Equipment, pk=product_id)  # ← ИСПРАВЛЕНО!
        form = SubtractQuantityForm()

    return render(request, 'accounting/subtract_quantity.html', {
        'form': form,
        'product': product,
        'previous_url': request.META.get('HTTP_REFERER', '/'),
    })

# Добавление количества
@transaction.atomic
def add_quantity_view(request, product_id):
    if request.method == 'POST':
        form = SubtractQuantityForm(request.POST)
        if form.is_valid():
            amount = form.cleaned_data['amount']

            # Блокируем строку на время транзакции
            add_product = Equipment.objects.select_for_update().get(pk=product_id)


            add_product.quantity += amount
            add_product.save()

            messages.success(request, f"Успешно добавлено {amount} шт.")

            # Возврат на страницу, откуда пришли
            next_url = request.GET.get('next', 'board_android_list')
            return redirect(next_url)
    else:
        add_product = get_object_or_404(Equipment, pk=product_id)
        form = SubtractQuantityForm()

    return render(request, 'accounting/add_quantity.html', {
        'form': form,
        'add_product': add_product,
        'previous_url': request.META.get('HTTP_REFERER', '/'),
    })